import datetime,subprocess,json
from pathlib import Path
from openpyxl import load_workbook
import pandas as pd

def log_subprocess(result: subprocess.CompletedProcess, log_dir: Path, id: str):
    """
    Collects logs produced by python subprocess and puts them in a text file for easy viewing
    Params:
        result:                     subprocess.CompletedProcess object that can be logged, result of running subprocess
        log_dir:                    path to the directory to store logs in
        id:                         identifier string for this process being logged
    """
    # get timestamp
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # make sure log_dir exist
    log_dir.mkdir(parents=True,exist_ok=True)

    # path to log file
    log_file = log_dir / "subprocess_log.jsonl"

    # dict of values to store
    data = {
        "id": id,
        "log_ts": timestamp,
        "stdout": result.stdout,
        "stderr": result.stderr
    }

    with open(log_file, "a") as f:
        f.write(json.dumps(data) + "\n")

def generate_template(file: Path = "template.xlsx"):
    """
    generates a templeate xlsx file for iputting m/z and rt values
    """

    header1 = "Template file for gcms automatic peak picking/integration, please ONLY fill in appropriate values and feel free to leave case/control empty if need be"
    header2 = "molecule = id of this moleucke, mz = ion to measure, rt = peak retention time case/control = list of sample names in each group"

    df = pd.DataFrame(columns=["molecule","mz","rt","case","control"])
    df.to_excel(file,index=False,startrow=3,startcol=1)

    wb = load_workbook(file)
    ws = wb.active

    ws["A1"] = header1
    ws["A2"] = header2

    wb.save(file)
    