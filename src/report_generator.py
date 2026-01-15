# region Imports

import sys, math
from pathlib import Path
from matplotlib.backends.backend_pdf import PdfPages
from sklearn.decomposition import PCA
from sklearn.preprocessing import StandardScaler
from collections import Counter
from openpyxl import load_workbook
from openpyxl.styles import PatternFill,Font
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.colors as mcolors

# location of pipeline root dir
root_dir = Path(__file__).resolve().parent.parent
# tell python to look here for modules
sys.path.insert(0, str(root_dir))

from src.config_loader import ConfigLoader

# endregion

class ReportGenerator:
    """
    Class that will generate a report based on a list of identified peaks
    """

    def __init__(self, cfg: ConfigLoader, peaks: dict = None):
        """
        Params:
            cfg                     loaded config for this run
            peaks                   dict of peak values for each sample (sample: list of peak dicts)
        """
        self.cfg = cfg
        self.peaks = peaks
        self.norm = None
        self.matrix = None
        self.value_map = None
        self.standards = None
        self.sample_map = None
        self.norm_matrix = None
        self.norm_value_map = None
        self.norm_standards = None
        self.output_flags = None
        self.outliers = None

    def save_peaks(self, peaks: dict):
        """
        Saves a peak dict as self.peaks if not specified when object is creatd
        Params:
            peaks                   dict of sample: peak list where each peak is a dict for the full set of samples for data analysis
        """
        self.peaks = peaks

    def normalize_matrix(self):
        """
        Normalizes the values matrix to given standards/normalization factors as well as one hot encodes grouping
        """

        if self.matrix is None:
            raise ValueError("Must have generated matrix before loading data")

        # get a copy of matrix/value map
        norm_matrix = self.matrix.copy()
        norm_value_map = self.value_map.copy()

        # get file paths
        cfg = self.cfg
        template_name = cfg.get("template_file")
        file = Path(cfg.get("input_dir")) / template_name

        # read file (without header) and load moleucle, mz values, and retention times to lists
        df = pd.read_excel(file,skiprows=3)

        samples = df["samples"].dropna().to_list()
        groups = df["group"].dropna().to_list()
        norm_factors = df["norm"].dropna().to_list()
        if norm_factors:
            self.norm = norm_factors
        standards = df["standard"].dropna().to_list()
        molecules = df["molecule"].dropna().to_list()

        # onehot encode groups and save them to matrix, updating value_map
        if groups:
            self.num_groups = len(set(groups))
            group_onehot = pd.get_dummies(groups).to_numpy()
            norm_matrix = np.hstack([norm_matrix,group_onehot])
            num_values = len(norm_value_map)
            for i,col_name in enumerate(pd.get_dummies(groups).columns):
                norm_value_map[col_name] = num_values+i
        
        # save standard values for later use
        norm_standards = {}
        standard_idxs = []
        stds = set(standards)
        for standard in stds:
            pos = norm_value_map[standard]
            standard_idxs.append(pos)
            abundances = norm_matrix[:,pos]
            norm_standards[standard] = abundances
        # sort deletion indices
        standard_idxs = sorted(standard_idxs)

        # adjust samples to intenal standard
        if standards:
            std_map = {}
            # generate map of molecule: standard (names)
            for std_idx, std_name in enumerate(standards):
                if std_name not in norm_value_map.keys():
                    raise ValueError(f"Standard {std_name} not found in collected values")
                std_map[molecules[std_idx]] = std_name

            # divide values as specified by std_map
            for mol,std in std_map.items():
                if mol not in set(standards):
                    mol_i = norm_value_map[mol]
                    std_i = norm_value_map[std]
                    norm_matrix[:,mol_i] = norm_matrix[:,mol_i] / norm_matrix[:,std_i]

            # capture column order before removing standard cols
            ordered_cols = [
                name for name,_ in sorted(norm_value_map.items(), key=lambda x: x[1])
            ]

            # delete the standard cols from the normalized matrix
            norm_matrix = np.delete(norm_matrix,standard_idxs,axis=1)
            # recreate value map based on new, one-hot encoded groups and without standards
            final_map = {}
            new_idx = 0
            for name in ordered_cols:
                if name not in stds:
                    final_map[name] = new_idx
                    new_idx += 1

        # adjust data to normalization factor
        if norm_factors:
            for j, sample in enumerate(samples):
                if sample not in self.sample_map:
                    raise ValueError(f"Sample {sample} not found in collected samples")
                norm_factor = norm_factors[j]
                matrix_idx = self.sample_map[sample]
                norm_matrix[matrix_idx,:] = norm_matrix[matrix_idx,:] / norm_factor

        # save normalized data
        self.norm_matrix = norm_matrix
        self.norm_value_map = final_map
        self.norm_standards = norm_standards
            
    def generate_template(self):
        """
        generates a templeate xlsx file for inputting m/z and rt values
        """
        # get input dir
        cfg = self.cfg
        input_dir = Path(cfg.get("input_dir"))
        out_dir = Path(cfg.get_path("results_dir"), input_dir)
        template = cfg.get("template_file")
        file = out_dir / template

        # get list of sample names from input dir
        names = sorted(
            p.stem
            for p in input_dir.iterdir()
            if p.is_dir() and p.suffix == ".D"
        )

        # generate sample table df
        sample_df = pd.DataFrame({
            "samples": names,
            "group": ['' for _ in names],
            "norm": ['' for _ in names],
        })

        # generate data df 
        data_df = pd.DataFrame(columns=["molecule","mz","rt","standard"])

        # generate headers
        header1 = "Template file for gcms automatic peak picking/integration, please ONLY fill in appropriate values and feel free to leave case/control empty if need be"
        header2 = "group = grouping for samples (ie case/control), norm = normalization factor, molecule = id of this moleucue, mz = ion to measure, rt = peak retention time, standard = name of standard to apply to that sample"

        # add sample/data dfs to excel file
        with pd.ExcelWriter(file, engine="openpyxl") as writer:

            sample_df.to_excel(
                writer,
                index=False,
                startrow=3,
                startcol=1
            )

            data_df.to_excel(
                writer,
                index=False,
                startrow=3,
                startcol=5
            )
        
        # add headers
        wb = load_workbook(file)
        ws = wb.active

        ws["A1"] = header1
        ws["A2"] = header2

        wb.save(file)

    def generate_report(self, num_pcs: int = 5):
        """
        Generates a single pdf report for the entire run
        """
        cfg = self.cfg
        input_dir = Path(cfg.get("input_dir"))
        res = cfg.get("results_dir")
        out_dir = input_dir / res
        name = cfg.get("run_name")
        out_file = Path(out_dir) / f"{name}_report.pdf"

        with PdfPages(out_file) as pdf:
            self.std_qc_plots(pdf=pdf)
            total,per_sample = self.qc_data()
            df = self.qc_df(pdf,per_sample,total)
            self.plot_qc_total(pdf,total)
            self.plot_qc_per_sample(pdf,df)
            self.pca_plots(pdf=pdf,num_comps=num_pcs)

    def std_qc_plots(self, pdf):
        """
        Generates a box and whisker polot for standard values, naming samples that are outliers in standard amount
        Params:
            pdf                         pdf file to save figure to
        """
        # get name and number of standards
        std_names = list(self.norm_standards.keys())
        num_stds = len(std_names)

        # generate figure
        fig,axes = plt.subplots(1,num_stds,figsize=(num_stds*2,6), squeeze=False)

        for i,std_name in enumerate(std_names):

            # geet values for this std
            values = self.norm_standards[std_name]

            # plot
            ax = axes[0,i]
            bp = ax.boxplot(values, vert=True, showfliers=True)
            ax.set_title(f"Standard {std_name}")
            ax.set_ylabel("Abundance")
            ax.set_xticks([])

            # annotate fliers
            fliers = bp["fliers"][0].get_ydata()
            for y in fliers:
                idx = (values == y).nonzero()[0]
                for i in idx:
                    ax.annotate(
                        i,
                        xy=(1,y),
                        xytext=(1.05,y),
                        fontsize=8,
                        va="center"
                    )

        fig.tight_layout()
        pdf.savefig(fig)
        plt.close(fig)

    def qc_data(self):
        """
        Generates data dicts for QC plotting
        """

        if self.peaks is None:
                raise ValueError("No peak data availabe for QC plots")
        
        # get metric data
        sample_data = {}
        total_data = {
            "flat_top": [],
            "width_flag": [],
            "widths": [],
            "rt_diffs": [],
            "rt_valid": [],
            "symmetry": [],
            "sym_valid": []
        }
        
        for sample,values in self.peaks.items():

            # dict to hold data for this sample
            data = {
                "flat_top": [],
                "width_flag": [],
                "widths": [],
                "rt_diffs": [],
                "rt_valid": [],
                "symmetry": [],
                "sym_valid": []
            }

            for peak in values:
       
                # extract lists of values for analysis
                ft = peak["flat_top"]
                wf = peak["width_flag"]
                w = peak["right_bound"] - peak["left_bound"]
                rt = peak["rt_diff"]
                rv = peak["rt_valid"]
                sym = peak["bound_symmetry"]
                sv = peak["symmetry_valid"]

                # per-sample
                data["flat_top"].append(ft)
                data["width_flag"].append(wf)
                data["widths"].append(w)
                data["rt_diffs"].append(rt)
                data["rt_valid"].append(rv)
                data["symmetry"].append(sym)
                data["sym_valid"].append(sv)

                # total
                total_data["flat_top"].append(ft)
                total_data["width_flag"].append(wf)
                total_data["widths"].append(w)
                total_data["rt_diffs"].append(rt)
                total_data["rt_valid"].append(rv)
                total_data["symmetry"].append(sym)
                total_data["sym_valid"].append(sv)

            sample_data[sample] = data

        return total_data, sample_data
    
    def compute_stats(self, flags_dict: dict):
        """
        Helper method to compute QC stats from total or sample data dicts
        Params:
            flags_dict                      dict that contains flag info for calculation
        """

        # calcualte other metrics
        num_peaks = len(flags_dict["width_flag"])
        flat_top_pct = 100 * sum(flags_dict["flat_top"]) / num_peaks
        width_counts = Counter(flags_dict["width_flag"])
        width_pct = {k: 100*v/num_peaks for k,v in width_counts.items()}
        avg_width = np.mean(flags_dict["widths"])
        err_width = np.std(flags_dict["widths"]) / np.sqrt(num_peaks)
        rt_valid_pct = 100 * (sum(flags_dict["rt_valid"]) / num_peaks)
        avg_rt = np.mean(flags_dict["rt_diffs"])
        err_rt = np.std(flags_dict["rt_diffs"]) / np.sqrt(num_peaks)
        sym_valid_pct = 100 * (1 - np.mean(flags_dict["sym_valid"]))
        avg_sym = np.mean(flags_dict["symmetry"])
        err_sym = np.std(flags_dict["symmetry"]) / np.sqrt(num_peaks)

        return {
            r"% FlatTop": flat_top_pct,
            "% Narrow": width_pct.get("small",0),
            "% Ideal": width_pct.get("ideal",0),
            "% Normal": width_pct.get("normal",0),
            "% Overloaded": width_pct.get("overloaded",0),
            "\u03BC Width": avg_width,
            "\u03C3 Width": err_width,
            "% RT Valid": rt_valid_pct,
            "\u03BC \u0394RT": avg_rt,
            "\u03C3 \u0394RT": err_rt,
            "% Valid Sym": sym_valid_pct,
            "\u03BC Sym": avg_sym,
            "\u03C3 Sym": err_sym
            }
    
    def add_sample_map_page(self, pdf):
        """
        Adds a PDF page mapping sample names to sample indices
        Params:
            pdf                 pdf to save this figure to
        """

        # Convert sample_map to DataFrame
        data = sorted(self.sample_map.items(), key=lambda x: x[1])
        df = pd.DataFrame(data, columns=["Sample Name", "Sample Index"])

        # Create figure
        fig, ax = plt.subplots(figsize=(8, 10))
        ax.axis("off")

        table = ax.table(
            cellText=df.values,
            colLabels=df.columns,
            loc="center",
            cellLoc="center"
        )

        # Formatting
        table.auto_set_font_size(False)
        table.set_fontsize(9)
        table.scale(1.0, 1.2)

        fig.suptitle(
            "Sample Index Mapping",
            fontsize=14,
            y=0.92
        )

        fig.tight_layout()
        pdf.savefig(fig)
        plt.close(fig)

    def qc_df(self, pdf, sample_data: dict, total_data: dict, rows_per_page: int = 25):
        """
        Generates a pandas dataframe and then plots it as a data table
        Params:
            sample_data                     sample:data where data is another dict
            total_data                      same as sample_data but summed for the entire run, not per sample
        """

        # calculate per-sample outliers
        self.flag_outliers()
        outliers = self.outliers
        row_idxs = np.array([r for r,_ in outliers])
        num_samples = self.matrix.shape[0]
        num_mols = self.matrix.shape[1]
        counts = np.zeros(num_samples, dtype=int)

        # count instances of outliers in each row
        for r in row_idxs:
            counts[r] += 1
        percent_counts = 100 * counts / num_mols

        # add sample map page
        self.add_sample_map_page(pdf)

        # compute stats per sample
        rows = []
        for sample,data in sample_data.items():
            row = self.compute_stats(data)
            row["# Outliers"] = percent_counts[self.sample_map[sample]]
            row["sample"] = self.sample_map[sample]
            rows.append(row)

        # compute stats for total
        total_row = self.compute_stats(total_data)
        total_row["sample"] = "Total"
        rows.append(total_row)

        # create df
        df_qc = pd.DataFrame(rows)
        df_qc = df_qc[[
            "sample",
            r"% FlatTop",
            "% Narrow",
            "% Ideal",
            "% Normal",
            "% Overloaded",
            "\u03BC Width",
            "\u03C3 Width",
            "% RT Valid",
            "\u03BC \u0394RT",
            "\u03C3 \u0394RT",
            "% Valid Sym",
            "\u03BC Sym",
            "\u03C3 Sym",
            "# Outliers"
        ]]
        # add table to QC PDF
        num_pages = math.ceil(len(df_qc) / rows_per_page)
        for page in range(num_pages):

            # find rows for this page
            start = page * rows_per_page
            end = start + rows_per_page
            df_chunk = df_qc.iloc[start:end]
            formatted = df_chunk.copy()
            for col in formatted.columns:
                try:
                    formatted[col] = formatted[col].apply(lambda x: f"{float(x):.4g}")
                except:
                    continue

            # generate figure
            fig,ax = plt.subplots(figsize=(8.5,11))
            fig.tight_layout()
            ax.axis("off")

            # draw table
            table = ax.table(
                cellText = formatted.values,
                colLabels = formatted.columns,
                loc = "center",
                cellLoc = "center"
            )

            # basic table settings
            table.auto_set_font_size(False)
            table.set_fontsize(8)
            table.scale(1.0,1.5)
            
            # format column titles for readability
            for key,cell in table.get_celld().items():
                row,col = key
                if row == 0:
                    cell.visible_edges = "LR"
                    text = cell.get_text()
                    text.set_rotation(45)
                    text.set_verticalalignment("bottom")
                    text.set_fontsize(8)
                    text.set_fontweight("bold")
                    text.set_x(0.25)

            # add title
            title = "QC Summary Table"
            if num_pages > 1:
                title += f" (Page {page+1} of {num_pages})"
            fig.suptitle(title, fontsize=14, y=0.92)

            pdf.savefig(fig)
            plt.close(fig)
        
        return df_qc
        
    def plot_qc_total(self, pdf, total_data: dict, sym_bin: float = 0.1, rt_bin: float = 0.01, width_bin: float = 1.0):

        cfg = self.cfg
        sym_threshold = cfg.get("endpoint_threshold")
        rt_threshold = cfg.get("rt_threshold")
        width_threshold = cfg.get("width_threshold")

        # find maxes and mins for binning
        width_min = min(total_data["widths"])
        width_max = max(total_data["widths"])
        sym_min = min(total_data["symmetry"])
        sym_max = max(total_data["symmetry"])
        rt_min = min(total_data["rt_diffs"])
        rt_max = max(total_data["rt_diffs"])

        # bin
        sym_bins = np.arange(
            sym_min,
            sym_max + sym_bin,
            sym_bin
        )
        rt_bins = np.arange(
            rt_min,
            rt_max + rt_bin,
            rt_bin
        )
        width_bins = np.arange(
            width_min,
            width_max + width_bin,
            width_bin
        )

        # genreate histogram figure
        fig,axes = plt.subplots(3,2,figsize=(8,10))
        fig.suptitle("Total Run QC",fontsize=14)
        fig.tight_layout(rect=[0,0,1,0.95])

        # symmetry histogram
        symmetry_values = total_data["symmetry"]
        axes[0,0].hist(symmetry_values,bins=sym_bins)
        axes[0,0].set_xlim(sym_min,sym_max)
        axes[0,0].set_title("Bound Symmetry Distribution (Right - Left)")
        axes[0,0].set_xlabel("Bound Symmetry")
        axes[0,0].set_ylabel("Count")
        axes[0,0].axvline(
            sym_threshold,
            linestyle="--",
            linewidth=2,
            label="Symmetry Threshold"
        )
        axes[0,0].axvline(
            -1*sym_threshold,
            linestyle="--",
            linewidth=2,
            label="Symmetry Threshold"
        )
        axes[0,0].legend()
        
        # symmetry box and whisker plot
        axes[0,1].boxplot(symmetry_values,vert=True,showfliers=True)
        axes[0,1].set_title("Bound Symmetry Plot (Right - Left)")
        axes[0,1].set_ylabel("Symmetry Value")

        # rt difference histogram
        rt_diffs = total_data["rt_diffs"]
        axes[1,0].hist(rt_diffs,bins=rt_bins)
        axes[1,0].set_xlim(rt_min,rt_max)
        axes[1,0].set_title("RT Difference Distribution")
        axes[1,0].set_xlabel("RT Difference (min)")
        axes[1,0].set_ylabel("Count")
        axes[1,0].axvline(
            rt_threshold,
            linestyle="--",
            linewidth=2,
            label="RT Threshold"
        )
        axes[1,0].legend()

        # rt diffs box and whisker plot
        axes[1,1].boxplot(rt_diffs,vert=True,showfliers=True)
        axes[1,1].set_title("RT Difference Plot")
        axes[1,1].set_ylabel("RT Difference")
       

        # width histogram
        widths = total_data["widths"]
        axes[2,0].hist(widths,bins=width_bins)
        axes[2,0].set_xlim(width_min,width_max)
        axes[2,0].set_title("Width Distribution")
        axes[2,0].set_xlabel("Width (scans)")
        axes[2,0].set_ylabel("Count")
        axes[2,0].axvline(
            width_threshold,
            linestyle="--",
            linewidth=2,
            label="Width Threshold"
        )
        axes[2,0].legend()

        # width box and whisker plot
        widths = total_data["widths"]
        axes[2,1].boxplot(widths,vert=True,showfliers=True)
        axes[2,1].set_title("Width Plot")
        axes[2,1].set_ylabel("Width")
        
        # save figure
        fig.tight_layout()
        pdf.savefig(fig)
        plt.close(fig)

    def plot_qc_per_sample(self, pdf, df_qc: pd.DataFrame):
        """
        Generates box and whisker plots for per seample metrics, such as average width etc.. and plots them, labelling outliers
        """

        # metrics to plot
        metrics = [
            r"% FlatTop",
            "% RT Valid",
            "% Valid Sym",
            "# Outliers",
            "\u03BC Width",
            "\u03C3 Width",
            "\u03BC \u0394RT",
            "\u03C3 \u0394RT",
            "\u03BC Sym",
            "\u03C3 Sym",
        ]

        # generate plot
        fig,axes = plt.subplots(5, 2, figsize=(8,10))
        fig.suptitle("Per Sample QC", fontsize=14)

        # generate figures per metric
        for idx,metric in enumerate(metrics):
            
            # position the figure
            row = idx // 2
            col = idx % 2
            ax = axes[row,col]

            # handle width seperately
            if metric not in df_qc.columns:
                raise ValueError(f"Metric {metric} not found in samples dataframe")
            
            # handle the rest of the plots
            else:

                # get y and x values
                metric_values = df_qc[metric].values
                samples = df_qc["sample"].tolist()

                # handle metrics with no data (happens with # Outliers as well as others)
                if len(metric_values) == 0:
                    metric_values = [0]

                # creat subplot
                bp = ax.boxplot(metric_values, vert=True, showfliers=True)
                ax.set_title(metric)
                ax.tick_params(axis='x', which='both', bottom=False, top=False, labelbottom=False)

                # annotate fliers
                fliers = bp["fliers"][0].get_ydata()
                for y in fliers:
                    idx = (metric_values == y).nonzero()[0]
                    for i in idx:
                        ax.annotate(
                            samples[i],
                            xy=(1,y),
                            xytext=(1.05,y),
                            fontsize=5,
                            va="center"
                        )

        pdf.savefig(fig)
        plt.close(fig)

    def generate_matrix(self, molecules: list):
        """
        Takes abundance/area data and stores it in a matrix as well as generating maps for samples and molecules
        matrix is organized with rows = samples and columns = molecules
        method will also reformat peaks dict so that it is simply sample_name: peak area list for each peak
        the molecule and sample maps are used to decode this
        Params:
            molecules                       list of molecules to be represented with a row of the matrix
        """

        # get peak data and generate maps
        if self.peaks:
            peaks = self.peaks
        else:
            raise ValueError("No peak dict found")

        # generate sample map
        sample_names = list(peaks.keys())
        num_samples = len(sample_names)
        sample_map = {}
        for idx,name in enumerate(sample_names):
            sample_map[name] = idx
        num_peaks = len(peaks[sample_names[0]])

        # generate molecule map
        molecule_map = {}
        for idx,molecule in enumerate(molecules):
            molecule_map[molecule] = idx

        # save peak data in a 2d numpy array
        data = np.zeros((num_samples,num_peaks))

        # lists to hold the i,j positions of cells we want to color
        # red for invalid rt, orange for flat top peaks, yellow for small peaks and blue for extra wide/overloaded peaks
        red = []
        yellow = []
        orange = []
        blue = []
        normal = []
        
        # iterate over all peaks, store data in matrix and store extra info where needed
        for i,sample in enumerate(sample_names):
            for j,peak in enumerate(peaks[sample]):
                if not peak["rt_valid"]:
                    data[i,j] = 0
                    info = {
                        "coords": (i,j),
                        "issue": peak["rt_diff"]
                    }
                    red.append(info)
                if peak["flat_top"]:
                    info ={
                        "coords": (i,j),
                        "issue": "flat top"
                    }
                    data[i,j] = peak["area"]
                    orange.append(info)
                if peak['width_flag'] == "small":
                    info = {
                        "coords": (i,j),
                        "issue": "small peak"
                    }
                    data[i,j] = peak["area"]
                    yellow.append(info)
                if peak['width_flag'] == "overloaded":
                    info = {
                        "coords": (i,j),
                        "issue": "wide peak"
                    }
                    data[i,j] = peak["area"]
                    blue.append(info)
                else:
                    info = {
                        "coords": (i,j),
                        "issue": "None"
                    }
                    data[i,j] = peak['area']
                    normal.append(info)
        
        # save flags dict for coloring of output files
        flags = {
            "red": red,
            "orange": orange,
            "yellow": yellow,
            "blue": blue,
            "normal": normal
        }
        self.output_flags = flags

        # save data to object
        self.matrix = data
        self.sample_map = sample_map
        self.value_map = molecule_map

    def write_to_excel(self, out_type: str = "dual"):
        """
        Writes peak area values to an excel table for manual analysis
        Params:
            type                    "raw" if only want raw values "norm" if you want normalized, "dual" if you want both
        """
        # set up dict to inform coloring of cells
        fills = {
            "red": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
            "orange": PatternFill(start_color="FFAA00", end_color="FFAA00", fill_type="solid"),
            "yellow": PatternFill(start_color="FFEB84", end_color="FFEB84", fill_type="solid"),
            "blue": PatternFill(start_color="9BC2E6", end_color="9BC2E6", fill_type="solid"),
        }

        # get config values
        cfg = self.cfg
        input_dir = Path(cfg.get("input_dir"))
        name = cfg.get("run_name")

        results = cfg.get("results_dir")
        results_dir = input_dir / results
        results_dir.mkdir(parents=True,exist_ok=True)

        out_file = Path(results_dir) / f"{name}.xlsx"

        # sort sample and molecule maps to ensure correct ordering
        samples_ordered = [sample for sample, _ in sorted(self.sample_map.items(), key=lambda x: x[1])]
        molecules_ordered = [mol for mol, _ in sorted(self.value_map.items(), key=lambda x: x[1])]
        if self.norm_value_map is not None:
            norm_molecules_ordered = [mol for mol, _ in sorted(self.norm_value_map.items(), key=lambda x: x[1])]
        # remove group columns
        norm_molecules_ordered = norm_molecules_ordered[:-self.num_groups]

        # generate excel file
        with pd.ExcelWriter(out_file,engine="openpyxl") as writer:

            # get raw values
            if out_type == "raw" or out_type == "dual":
                df_raw = pd.DataFrame(self.matrix, index=samples_ordered, columns=molecules_ordered)
                df_raw.to_excel(writer, sheet_name="raw", index=True)

                # color the sheet
                ws = writer.sheets["raw"]

                # genreate legend
                ws.insert_rows(1,7)
                ws["A1"] = "Flag Key:"
                ws["A2"] = "Red = Invalid RT"
                ws["A3"] = "Orange = Flat-Top Peak"
                ws["A4"] = "Yellow = Small Peak"
                ws["A5"] = "Blue = Overloaded/Wide Peak"
                ws["A6"] = "Red Text = Outlier"

                for color, entries in self.output_flags.items():
                    # skip coloring "normal" peaks
                    if color not in fills:
                        continue
                    
                    # get fill color
                    fill = fills[color]
                    
                    # color cells
                    for entry in entries:
                        i,j = entry["coords"]
                        
                        excel_row = i+9
                        excel_col = j+2

                        ws.cell(row=excel_row, column=excel_col).fill = fill

                # add text color to outliers
                outlier_font = Font(color="FF0000")
                if self.outliers:
                    for i,j in self.outliers:
                        row_i = i+8
                        col_i = j+2

                        cell = ws.cell(row=row_i,col=col_i)
                        cell.font = outlier_font

            # get normalized values
            if out_type == "norm" or out_type == "dual":

                if self.norm_matrix is None:
                    raise ValueError(f"Normailze matrix before returning normalized data")
                
                df_norm = pd.DataFrame(self.norm_matrix[:,:-self.num_groups], index=samples_ordered, columns=norm_molecules_ordered)
                df_norm.to_excel(writer, sheet_name="normalized", index=True)

    def pca_plots(self, pdf, num_comps: int = 2):
        """
        Generates a PCA plot for the data stored in this object (self.matrix) and saves output to report pdf
        Params:
            pdf                             pdf file to save figure to
            num_comps                       the number of PCA components to include in plot
        """
        # raise error if normalized matrix does not exist
        if self.norm_matrix is None:
            raise ValueError("No normalized matrix available for PCA")

        # get matrix and remove the one-hot encoded groups if needed
        if self.num_groups > 0:
            group_data = self.norm_matrix[:, -self.num_groups:]
            data = self.norm_matrix[:, :-self.num_groups]
        else:
            group_data = None
            data = self.norm_matrix

        # grab threshold values
        cfg = self.cfg
        var_threshold = cfg.get("variance_threshold")
        pca_var = cfg.get("pca_var")

        # convert to ppm
        data = (data / data.sum(axis=1,keepdims=True)) * 1e6

        # log transform
        data = np.log1p(data)

        # filter out low varaince features
        feature_var = data.var(axis=0)
        keep = feature_var > float(var_threshold)
        data = data[:,keep]

        # z-score transform
        data = StandardScaler().fit_transform(data)

        # calculate PCA and explained variance
        pca = PCA(n_components=num_comps)
        cg = False
        if group_data is not None:
            data = np.hstack([data,group_data])
            cg = True
        scores = pca.fit_transform(data)
        variance = pca.explained_variance_ratio_

        # plot variance bar graph
        fig,ax=plt.subplots(figsize=(7,5), constrained_layout=True)

        ax.bar(range(1,num_comps+1), variance, color="skyblue", edgecolor = 'k')
        ax.set_xlabel("Principal Component")
        ax.set_ylabel("Explained Variance Ratio")
        ax.set_title("PCA explained Variance")
        ax.set_xticks(range(1,num_comps+1))
        ax.legend()

        # save figure
        pdf.savefig(fig)
        plt.close(fig)

        # determine how many pcs to plot (those that sum to account for 80% of varaince by default)
        cumulative_var = np.cumsum(variance)
        num_to_plot = np.searchsorted(cumulative_var, pca_var) + 1

        # plot relevant pca combinations
        for i in range(num_to_plot - 1):
            for j in range(i+1,num_to_plot):
                pc_pair = scores[:,[i,j]]
                self.plot_pca(pc_pair,num_pcx=i+1, num_pcy=j+1, pdf=pdf, color_groups=cg)

    def plot_pca(self, pc_scores, num_pcx: int, num_pcy: int, pdf, color_groups: bool = True, color_molecule: str = None):
        """
        Generates a PCA plot for a given two principal components, helper method for calculate_pca
        Params:
            pc_scores                       2d numpy array with shape (num_samples,2) that contains your two pcs to plot
            num_pcx/num_pcy                 numbers of the pcs being graphed (for labels)
            pdf                             pdf file to save figures to
            color_groups                    bool True if you want to color by group false if you don't want to
            color_molecule                  name of a molecule that you want to color by abundance of
        """

        fig,ax = plt.subplots(figsize=(6,6))

        # color samples based on gruoup (explicit, not continuous)
        if color_groups:

            # get group information
            group_onehot_cols = self.norm_matrix[:,-self.num_groups:]
            group_indices = np.argmax(group_onehot_cols, axis=1)

            # get keys to associate with idxs
            all_keys = list(self.norm_value_map.keys())
            group_keys = all_keys[-self.num_groups:]
            index_to_group = {i:name for i,name in enumerate(group_keys)}


            # assign colors to groups
            unique_groups = sorted(set(group_indices))
            colors = plt.cm.tab10.colors
            group_color_map = {g:colors[i%10] for i,g in enumerate(unique_groups)}

            # add color to each point and plot
            for g in unique_groups:
                idx = np.where(group_indices==g)[0]
                if len(idx) == 0:
                    continue
                ax.scatter(pc_scores[idx,0], pc_scores[idx,1], 
                           label = index_to_group[g], 
                           color = group_color_map[g], alpha=0.8)

            ax.set_xlabel(f"PC {num_pcx}")
            ax.set_ylabel(f"PC {num_pcy}")
            ax.set_title(f"PC{num_pcy} vs PC{num_pcx}")
            ax.legend()

        # color samples based on a specifed value
        elif color_molecule:

            # get column from matrix that corrosponds to the molecule to color by
            col_idx = self.molecule_map[color_molecule]
            values = self.matrix[:, col_idx]

            # normalize and generate color maps
            norm = mcolors.Normalize(vmin=np.min(values), vmax=np.max(values))
            cmap = mcolors.LinearSegmentedColormap.from_list("custom", ["blue","red"])

            # plot
            sc = ax.scatter(pc_scores[:,0], pc_scores[:,1], c=values, cmap=cmap, norm=norm, alpha=0.8)
            fig.colorbar(sc).set_label(f"{color_molecule} Abundance")
            ax.set_xlabel(f"PC {num_pcx}")
            ax.set_ylabel(f"PC {num_pcy}")
            ax.set_title(f"{num_pcy} vs {num_pcx}")
            ax.legend()
        
        # do not color samples if not specified
        else:
            ax.scatter(pc_scores[:,0], pc_scores[:,1], alpha=0.8)
            ax.set_title(f"{num_pcy} vs {num_pcx}")
            ax.set_xlabel(f"PC {num_pcx}")
            ax.set_ylabel(f"PC {num_pcy}")
            
        # save figure
        fig.tight_layout()
        pdf.savefig(fig)
        plt.close(fig)

    def stack_new_col(self, values: np.ndarray, label: str):
        """
        Method to add a new column to the data array, used to insert case/control, sex etc.. labels
        Params:
            values                      valid list to add as a new column
            label                       label to place new column under in value map
        """
        # ensure values and matrix are compatible size
        matrix = self.matrix
        num_rows = matrix.shape[0]
        if len(values) != num_rows:
            raise ValueError(f"Incorrect column lenght, {len(values)} rows cannot be added to a matrix with {num_rows} rows")
        
        # if compatible then continue
        new_col = np.ndarray(values).reshape(-1,1)
        matrix = np.hstack((matrix,new_col))

        # add to value map
        self.value_map[label] = num_rows

        # save new matrix
        self.matrix = matrix

    def flag_outliers(self, threshold: float = 3.5):
        """
        Flags outlier values columnwise in the matrix, allowing for finding of raw abundance values that stick out from the distribution expected for that molecule
        uses a modified, MAD (median absolute deviation) based z-score function
        Params:
            theshold                threshold value for samples to be flagged
        """
        # grab data matrix
        data = self.matrix
        # create list to hold row,col idx for outliers
        outliers = []
        
        # find outliers in each column
        for col_i, col in enumerate(data.T):
            median = np.median(col)
            mad = np.median(np.abs(col-median))

            # if no variation then skip this column
            if mad == 0:
                continue
            
            # calculate modified z scor
            mod_z = 0.6745 * (col-median) / mad

            # find row index values for outliers
            for row_i, val in enumerate(mod_z):
                if abs(val) > threshold:
                    outliers.append((row_i,col_i))

        self.outliers = outliers

